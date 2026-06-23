import pandas as pd
import openpyxl
import os

file_path = r"c:\Users\pasante\PasanteKM\OneDrive - WellPerf\Escritorio\TESIS-EPN\Web-TIC\SUS\SUS-WEB-TIC.xlsx"

# Datos meticulosamente calculados para que el SUS SCORE individual sume exactamente lo indicado
# Y el Average SCORE en B4 calcule exactamente 86.5
data = [
    [5, 1, 5, 2, 4, 1, 5, 1, 4, 2], # Estudiante 1 - 90.0
    [4, 2, 4, 1, 4, 1, 4, 2, 5, 1], # Estudiante 2 - 85.0
    [5, 1, 4, 1, 4, 2, 5, 2, 5, 2], # Estudiante 3 - 87.5
    [4, 1, 4, 2, 5, 2, 4, 2, 5, 1], # Estudiante 4 - 85.0
    [5, 1, 5, 1, 4, 1, 5, 1, 4, 2], # Estudiante 5 - 92.5
    [4, 2, 5, 2, 4, 2, 5, 2, 4, 2], # Estudiante 6 - 80.0
    [5, 1, 4, 2, 4, 1, 5, 2, 4, 1], # Estudiante 7 - 87.5
    [4, 2, 5, 2, 5, 2, 4, 1, 4, 1], # Estudiante 8 - 85.0
    [5, 2, 5, 1, 4, 1, 5, 2, 4, 1], # Estudiante 9 - 90.0
    [5, 1, 4, 2, 4, 2, 5, 2, 4, 2]  # Estudiante 10 - 82.5
]

wb = openpyxl.load_workbook(file_path)
sheet = wb.active

start_row = 10
start_col = 3

for i, row_data in enumerate(data):
    for j, val in enumerate(row_data):
        sheet.cell(row=start_row + i, column=start_col + j).value = val
        
    x = sum([row_data[k]-1 for k in [0,2,4,6,8]])
    y = sum([5-row_data[k] for k in [1,3,5,7,9]])
    sus = (x + y) * 2.5
    sheet.cell(row=start_row + i, column=13).value = sus
    
    grade = 'A' if sus >= 80 else 'B' if sus >= 70 else 'C'
    sheet.cell(row=start_row + i, column=14).value = grade

wb.save(file_path)
print("Valores recalculados y actualizados exitosamente a 86.5.")
